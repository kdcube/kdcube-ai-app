# SPDX-License-Identifier: MIT
# Copyright (c) 2025 Elena Viter

from __future__ import annotations

import datetime as dt
import io
import json
import re
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Body, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from kdcube_ai_app.apps.chat.api.resolvers import auth_without_pressure, get_pg_pool
from kdcube_ai_app.apps.chat.sdk.config import get_settings
from kdcube_ai_app.apps.chat.sdk.context.retrieval.ctx_rag import ContextRAGClient
from kdcube_ai_app.apps.chat.sdk.context.vector.conv_index import ConvIndex
from kdcube_ai_app.apps.chat.sdk.storage.conversation_store import ConversationStore
from kdcube_ai_app.infra.service_hub.inventory import ConfigRequest, ModelServiceBase, create_workflow_config

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except Exception as exc:
    openpyxl = None  # type: ignore[assignment]
    _OPENPYXL_ERROR = exc
else:
    _OPENPYXL_ERROR = None


router = APIRouter()


class TenantProjectItem(BaseModel):
    tenant: str
    project: str
    schema: str
    source: str = Field(default="control_plane")


class TenantProjectListResponse(BaseModel):
    items: List[TenantProjectItem]


class UsersListResponse(BaseModel):
    tenant: str
    project: str
    items: List[str]


class ConversationFetchRequest(BaseModel):
    turn_ids: Optional[List[str]] = Field(default=None, description="If present, fetch only these turns")
    materialize: bool = Field(default=True, description="Fetch payloads from store for UI-visible items")
    days: int = Field(default=365, ge=1, le=3650)


def _sanitize_name(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9_]", "_", name.replace("-", "_").replace(" ", "_"))


def _schema_for(tenant: str, project: str) -> str:
    schema_name = f"{_sanitize_name(tenant)}_{_sanitize_name(project)}"
    if not schema_name.startswith("kdcube_"):
        schema_name = f"kdcube_{schema_name}"
    return schema_name


async def _get_pg_pool_from_state():
    pool = getattr(router.state, "pg_pool", None)
    if pool:
        return pool
    return await get_pg_pool()


def _get_shared_components():
    base_ctx = getattr(router.state, "conversation_browser", None)
    if base_ctx:
        return base_ctx.model_service, base_ctx.store

    settings = get_settings()
    req = ConfigRequest(
        openai_api_key=settings.OPENAI_API_KEY,
        claude_api_key=settings.ANTHROPIC_API_KEY,
        selected_model=settings.DEFAULT_MODEL_LLM,
    )
    model_service = ModelServiceBase(create_workflow_config(req))
    store = ConversationStore(settings.STORAGE_PATH)
    return model_service, store


def _build_ctx(pg_pool, tenant: str, project: str) -> ContextRAGClient:
    conv_idx = ConvIndex(pool=pg_pool)
    conv_idx.schema = _schema_for(tenant, project)
    model_service, store = _get_shared_components()
    return ContextRAGClient(conv_idx=conv_idx, store=store, model_service=model_service)


def _json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=True, sort_keys=True, default=str)


def _parse_iso(value: Optional[str]) -> Optional[dt.datetime]:
    if not value:
        return None
    try:
        return dt.datetime.fromisoformat(value.replace("Z", "+00:00"))
    except Exception:
        return None


def _format_ts(value: Optional[str]) -> str:
    parsed = _parse_iso(value)
    if not parsed:
        return ""
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=dt.timezone.utc)
    return parsed.astimezone(dt.timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def _extract_payload(data: Any) -> Dict[str, Any]:
    if isinstance(data, dict) and isinstance(data.get("payload"), dict):
        return data.get("payload") or {}
    if isinstance(data, dict):
        return data
    return {}


def _clean_attachment_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(payload, dict):
        return {}
    cleaned = {}
    for key, value in payload.items():
        if key in {"base64", "bytes"}:
            continue
        cleaned[key] = value
    return cleaned


def _format_attachment(payload: Dict[str, Any]) -> str:
    cleaned = _clean_attachment_payload(payload)
    meta = cleaned.get("meta") if isinstance(cleaned.get("meta"), dict) else {}
    rn = cleaned.get("rn") or meta.get("rn") or ""
    filename = cleaned.get("filename") or cleaned.get("name") or ""
    mime = cleaned.get("mime") or cleaned.get("mime_type") or ""
    size = cleaned.get("size") or cleaned.get("size_bytes") or ""
    hosted_uri = cleaned.get("hosted_uri") or cleaned.get("path") or cleaned.get("source_path") or ""
    parts = []
    if filename:
        parts.append(f"filename={filename}")
    if mime:
        parts.append(f"mime={mime}")
    if size:
        parts.append(f"size={size}")
    if rn:
        parts.append(f"rn={rn}")
    if hosted_uri and not rn:
        parts.append(f"uri={hosted_uri}")
    return "; ".join(parts) if parts else _json_dumps(cleaned)


def _format_bot_artifact(artifact: Dict[str, Any]) -> Optional[str]:
    art_type = artifact.get("type") or ""
    payload = _extract_payload(artifact.get("data"))
    meta = payload.get("meta") if isinstance(payload.get("meta"), dict) else {}
    if art_type == "artifact:assistant.file":
        filename = payload.get("filename") or payload.get("name") or ""
        mime = payload.get("mime") or payload.get("mime_type") or ""
        size = payload.get("size") or payload.get("size_bytes") or ""
        rn = payload.get("rn") or meta.get("rn") or ""
        hosted_uri = payload.get("hosted_uri") or payload.get("path") or payload.get("source_path") or ""
        parts = [art_type]
        if filename:
            parts.append(f"filename={filename}")
        if mime:
            parts.append(f"mime={mime}")
        if size:
            parts.append(f"size={size}")
        if rn:
            parts.append(f"rn={rn}")
        if hosted_uri and not rn:
            parts.append(f"uri={hosted_uri}")
        return " | ".join(parts)

    text = payload.get("text") or payload.get("payload", {}).get("text")
    if isinstance(text, str) and text.strip():
        return f"{art_type}: {text.strip()}"

    title = payload.get("title") or meta.get("title") or ""
    if title:
        return f"{art_type}: {title}"
    return None


def _format_citations(artifact: Dict[str, Any]) -> List[str]:
    payload = _extract_payload(artifact.get("data"))
    items = payload.get("items") if isinstance(payload.get("items"), list) else []
    lines: List[str] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        sid = item.get("sid")
        title = item.get("title") or ""
        url = item.get("url") or ""
        mime = item.get("mime") or ""
        text = item.get("text") or ""
        parts = []
        if sid is not None:
            parts.append(f"[{sid}]")
        if title:
            parts.append(str(title))
        if url:
            parts.append(str(url))
        line = " ".join(parts).strip()
        extras = []
        if mime:
            extras.append(f"mime: {mime}")
        if text:
            extras.append(f"text: {text}")
        if extras:
            line = f"{line} {'; '.join(extras)}".strip()
        if line:
            lines.append(line)
    return lines


def _sheet_name(base: str, used: set[str]) -> str:
    cleaned = re.sub(r"[:\\/?*\\[\\]]", "_", base).strip() or "conversation"
    cleaned = re.sub(r"[\\x00-\\x1F\\x7F]", "_", cleaned)
    cleaned = cleaned[:31]
    candidate = cleaned
    idx = 2
    while candidate in used:
        suffix = f"_{idx}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        idx += 1
    used.add(candidate)
    return candidate


async def _get_turn_timings(
    ctx: ContextRAGClient,
    user_id: str,
    conversation_id: str,
    turn_ids: List[str],
    days: int,
) -> Dict[str, Dict[str, Optional[str]]]:
    results: Dict[str, Dict[str, Optional[str]]] = {}
    for tid in turn_ids:
        mat = await ctx.materialize_turn(
            turn_id=tid,
            scope="conversation",
            days=days,
            user_id=user_id,
            conversation_id=conversation_id,
            with_payload=True,
        )
        turn_log = mat.get("turn_log") or {}
        payload = _extract_payload(turn_log.get("payload"))
        tl = payload.get("turn_log") if isinstance(payload.get("turn_log"), dict) else payload
        start = _parse_iso(tl.get("started_at_iso") or tl.get("started_at"))
        end = _parse_iso(tl.get("ended_at_iso") or tl.get("ended_at"))
        duration = (end - start).total_seconds() if start and end else None
        results[tid] = {
            "started_at": start.isoformat() if start else None,
            "ended_at": end.isoformat() if end else None,
            "duration": None if duration is None else f"{duration:.2f}",
        }
    return results


def _build_turn_rows(turns: List[Dict[str, Any]], timings: Dict[str, Dict[str, Optional[str]]]) -> List[List[str]]:
    rows: List[List[str]] = []
    for turn in turns:
        tid = turn.get("turn_id") or ""
        artifacts = turn.get("artifacts") or []
        user_messages: List[str] = []
        user_attachments: List[Any] = []
        assistant_messages: List[str] = []
        followups: List[str] = []
        bot_artifacts: List[str] = []
        citations: List[str] = []

        for art in artifacts:
            art_type = art.get("type") or ""
            data = art.get("data")
            if art_type == "chat:user":
                payload = _extract_payload(data)
                text = (payload.get("text") or "").strip()
                if text:
                    user_messages.append(text)
                continue
            if art_type == "chat:assistant":
                payload = _extract_payload(data)
                text = (payload.get("text") or "").strip()
                if text:
                    assistant_messages.append(text)
                continue
            if art_type == "artifact:user.attachment":
                user_attachments.append(_extract_payload(data))
                continue
            if art_type == "artifact:conv.user_shortcuts":
                payload = _extract_payload(data)
                items = payload.get("items") or []
                if isinstance(items, list):
                    followups.extend([str(i) for i in items if i])
                continue
            if art_type == "artifact:solver.program.citables":
                citations.extend(_format_citations(art))
                continue

            if art_type.startswith("artifact:"):
                formatted = _format_bot_artifact(art)
                if formatted:
                    bot_artifacts.append(formatted)

        timing = timings.get(tid) or {}
        rows.append([
            str(tid),
            _format_ts(timing.get("started_at")),
            _format_ts(timing.get("ended_at")),
            "\n\n".join(user_messages),
            "\n".join(_format_attachment(a) for a in user_attachments) if user_attachments else "",
            "\n".join(bot_artifacts),
            "\n".join(citations),
            "\n\n".join(assistant_messages),
            "\n".join(dict.fromkeys(followups)),
            timing.get("duration") or "",
        ])
    return rows


@router.get("/tenant-projects", response_model=TenantProjectListResponse)
async def list_tenant_projects(
    session=Depends(auth_without_pressure()),
):
    pool = await _get_pg_pool_from_state()
    items: Dict[str, TenantProjectItem] = {}
    settings = get_settings()

    def _add(tenant: str, project: str, source: str):
        key = f"{tenant}:{project}"
        if key in items:
            return
        items[key] = TenantProjectItem(
            tenant=tenant,
            project=project,
            schema=_schema_for(tenant, project),
            source=source,
        )

    # Control plane sources
    async with pool.acquire() as conn:
        for table in ("plan_quota_policies", "application_budget_policies"):
            try:
                rows = await conn.fetch(
                    f"""
                    SELECT DISTINCT tenant, project
                    FROM kdcube_control_plane.{table}
                    WHERE tenant IS NOT NULL AND project IS NOT NULL
                    """
                )
            except Exception:
                rows = []
            for row in rows:
                _add(row["tenant"], row["project"], f"control_plane.{table}")

    # Always include defaults as a fallback
    if settings.TENANT and settings.PROJECT:
        _add(settings.TENANT, settings.PROJECT, "defaults")

    return TenantProjectListResponse(items=sorted(items.values(), key=lambda x: (x.tenant, x.project)))


@router.get("/{tenant}/{project}/users", response_model=UsersListResponse)
async def list_users(
    tenant: str,
    project: str,
    limit: int = Query(default=200, ge=1, le=2000),
    search: Optional[str] = Query(default=None),
    session=Depends(auth_without_pressure()),
):
    pool = await _get_pg_pool_from_state()
    schema = _schema_for(tenant, project)
    args: List[Any] = [limit]
    where_clause = ""
    if search:
        args.append(f"%{search}%")
        where_clause = f"AND user_id ILIKE ${len(args)}"

    q = f"""
        SELECT DISTINCT user_id
        FROM {schema}.conv_messages
        WHERE user_id IS NOT NULL
        {where_clause}
        ORDER BY user_id
        LIMIT $1
    """
    async with pool.acquire() as conn:
        try:
            rows = await conn.fetch(q, *args)
        except Exception as exc:
            raise HTTPException(status_code=404, detail=f"Conversation schema not found: {schema}") from exc

    return UsersListResponse(tenant=tenant, project=project, items=[r["user_id"] for r in rows])


@router.get("/{tenant}/{project}/{user_id}/conversations")
async def list_user_conversations(
    tenant: str,
    project: str,
    user_id: str,
    last_n: Optional[int] = Query(default=None, ge=1, le=500),
    started_after: Optional[str] = Query(default=None),
    days: int = Query(default=365, ge=1, le=3650),
    include_titles: bool = Query(default=True),
    session=Depends(auth_without_pressure()),
):
    pool = await _get_pg_pool_from_state()
    ctx = _build_ctx(pool, tenant, project)
    started_after_dt = None
    if started_after:
        try:
            started_after_dt = dt.datetime.fromisoformat(started_after.replace("Z", "+00:00"))
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Invalid started_after timestamp") from exc

    return await ctx.list_conversations(
        user_id=user_id,
        last_n=last_n,
        started_after=started_after_dt,
        days=days,
        include_titles=include_titles,
    )


@router.get("/{tenant}/{project}/{user_id}/conversations/{conversation_id}/details")
async def get_conversation_details(
    tenant: str,
    project: str,
    user_id: str,
    conversation_id: str,
    session=Depends(auth_without_pressure()),
):
    pool = await _get_pg_pool_from_state()
    ctx = _build_ctx(pool, tenant, project)
    return await ctx.get_conversation_details(user_id=user_id, conversation_id=conversation_id)


@router.post("/{tenant}/{project}/{user_id}/conversations/{conversation_id}/fetch")
async def fetch_conversation(
    tenant: str,
    project: str,
    user_id: str,
    conversation_id: str,
    req: ConversationFetchRequest = Body(...),
    session=Depends(auth_without_pressure()),
):
    pool = await _get_pg_pool_from_state()
    ctx = _build_ctx(pool, tenant, project)
    return await ctx.fetch_conversation_artifacts(
        user_id=user_id,
        conversation_id=conversation_id,
        turn_ids=req.turn_ids or None,
        materialize=bool(req.materialize),
        days=int(req.days),
    )


@router.get("/{tenant}/{project}/{user_id}/export.xlsx")
async def export_user_conversations(
    tenant: str,
    project: str,
    user_id: str,
    conversation_ids: Optional[str] = Query(default=None, description="Comma-separated conversation IDs"),
    days: int = Query(default=365, ge=1, le=3650),
    session=Depends(auth_without_pressure()),
):
    if openpyxl is None:
        raise HTTPException(status_code=500, detail=f"Excel export unavailable: {_OPENPYXL_ERROR}")

    pool = await _get_pg_pool_from_state()
    ctx = _build_ctx(pool, tenant, project)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used_names: set[str] = set()

    header = [
        "turn_id",
        "turn started at",
        "turn ended at",
        "user message",
        "user attachments (metadata)",
        "produced by bot artifacts",
        "citations",
        "assistant final response",
        "suggested followups",
        "time taken (seconds)",
    ]
    header_fill = PatternFill(start_color="2B3A67", end_color="2B3A67", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    selected_ids: Optional[set[str]] = None
    if conversation_ids:
        selected_ids = {cid.strip() for cid in conversation_ids.split(",") if cid.strip()}
    if selected_ids:
        items = [{"conversation_id": cid, "title": None} for cid in sorted(selected_ids)]
    else:
        conversations = await ctx.list_conversations(user_id=user_id, days=days, include_titles=True)
        items = conversations.get("items") or []

    for item in items:
        conversation_id = item.get("conversation_id")
        if not conversation_id:
            continue
        details = await ctx.get_conversation_details(user_id=user_id, conversation_id=conversation_id)
        fetch_data = await ctx.fetch_conversation_artifacts(
            user_id=user_id,
            conversation_id=conversation_id,
            materialize=True,
            days=days,
        )
        turns = details.get("turns") or []
        turn_ids = [t.get("turn_id") for t in turns if t.get("turn_id")]
        timings = await _get_turn_timings(ctx, user_id, conversation_id, turn_ids, days)

        artifacts_by_turn = {
            t.get("turn_id"): t.get("artifacts") or []
            for t in (fetch_data.get("turns") or [])
            if t.get("turn_id")
        }

        merged_turns = []
        for t in turns:
            tid = t.get("turn_id")
            merged_turns.append({
                "turn_id": tid,
                "artifacts": artifacts_by_turn.get(tid, []),
            })

        sheet_title = item.get("title") or conversation_id
        ws = wb.create_sheet(title=_sheet_name(str(sheet_title), used_names))

        ws.append(header)
        for idx, _ in enumerate(header, start=1):
            cell = ws.cell(row=1, column=idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        for row in _build_turn_rows(merged_turns, timings):
            ws.append(row)

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 26
        ws.column_dimensions["D"].width = 60
        ws.column_dimensions["E"].width = 60
        ws.column_dimensions["F"].width = 70
        ws.column_dimensions["G"].width = 70
        ws.column_dimensions["H"].width = 60
        ws.column_dimensions["I"].width = 40
        ws.column_dimensions["J"].width = 22
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    if not wb.sheetnames:
        wb.create_sheet(title="No Conversations")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"{user_id}_conversations.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
